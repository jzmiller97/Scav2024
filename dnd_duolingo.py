import random
from quiz_games import Quiz
import openpyxl
import sys
import re

def true_length(text):
    escape_regex = re.compile(r'\x1b\[[0-9;]*m')
    clean_text = escape_regex.sub('', text)
    return len(clean_text)

def string_hearts(num_hearts) -> str:
    hearts: str = '\u001b[31m❤\u001b[0m '  # ANSI escape codes for red heart
    return hearts * num_hearts

def string_box(text, width: int):
    gap = abs(width - true_length(text))
    l1 = "\u001b[1m┏" + "━" * (width + 1) + "┓\u001b[0m"
    l2 = "\u001b[1m┃ " + text + " " * gap + "┃\u001b[0m"
    l3 = "\u001b[1m┗" + "━" * (width + 1) + "┛\u001b[0m"
    return "\n".join((l1, l2, l3))

def read_spreadsheet_to_dict(filename: str) -> dict[str, str]:
    data_dict: dict[str, str] = {}
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            key = str(row[0])
            value = str(row[1])
            data_dict[key] = value
    except FileNotFoundError:
        print(f"File '{filename}' not found.")
    except Exception as e:
        print("An error occurred:", e)
    return data_dict

def read_spreadsheet_to_set(filename: str) -> set[str]:
    data_set: set[str] = set()
    workbook = openpyxl.load_workbook(filename)
    for sheet in workbook.sheetnames:
        current_sheet = workbook[sheet]
        for row in current_sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    data_set.add(str(cell.value))
    return data_set

def print_sword_greeting():
    sword: list[str] = [
        "                 />",
        " ()            //---------------------------------------------------------(",
        "(*)OXOXOXOXOXO(*>  Greetings, adventurer. Welcome to Duolingo & Dragons.   \\",
        " ()            \\\\-----------------------------------------------------------)",
        "                 \\>"
    ]
    print("\u001b[33m")  # ANSI escape code for yellow color
    for line in sword:
        print(line)
    print("\u001b[0m")  # Reset color6

class DnD(Quiz):
    hp: int
    questions: dict[str, str]
    praises: set[str]
    insults: set[str]
    motivations: set[str]
    streak: int
    hp_max: int
    
    def __init__(self, questions: dict["str", "str"], praises: set[str], 
                 insults: set[str], motivations: set[str], hp: int) -> None:
        self.questions = questions
        self.praises = praises
        self.insults = insults
        self.motivations = motivations
        self.hp = hp
        self.hp_max = hp
        self.streak = 0
    
    def ask_question(self) -> None:
        question = random.choice(list(self.questions.keys()))
        motivation = random.choice(list(self.motivations))
        answer = input(question + " " + motivation + " ")
        self.check_answer(question, answer)

    def check_answer(self, question, user_answer) -> None:
        correct_answer = self.questions[question]
        if user_answer.lower() == correct_answer.lower():
            self.streak += 1
            if self.streak % 5 == 0 and self.hp < self.max_hp:
                self.hp += 1
            print(random.choice(list(self.praises)))
        else:
            self.streak = 0
            self.hp += -1
            print(random.choice(list(self.insults)) + " The correct answer is:", correct_answer)
            
    def reset(self) -> None:
        self.streak = 0
        self.hp = self.hp_max
        
    def print_gamestate(self, width: int) -> None:
        inp: str = "Hit points remaining: " + string_hearts(self.hp)
        print(string_box(inp, width))
        print("Current streak: " + str(self.streak) + ".")
        

if __name__ == "__main__":
    print_sword_greeting()
    dnd_questions: dict[str, str] = read_spreadsheet_to_dict("dnd_questions.xlsx")
    dnd_motivations: set[str] = read_spreadsheet_to_set("dnd_motivations.xlsx")
    dnd_praises: set[str] = read_spreadsheet_to_set("dnd_praises.xlsx")
    dnd_insults: set[str] = read_spreadsheet_to_set("dnd_insults.xlsx")
    quiz: "DnD" = DnD(dnd_questions, dnd_praises, dnd_insults, dnd_motivations, int(sys.argv[1]))
    width = true_length("Hit points remaining: " + string_hearts(quiz.hp)) + 2
    while True:
        quiz.print_gamestate(width)
        quiz.ask_question()
        if quiz.hp <= 0:
            play_again: str = input("Re-roll initiative? (yes/no): ")
            if play_again.lower() == "yes":
                quiz.reset()
            else:
                break